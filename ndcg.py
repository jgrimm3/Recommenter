import os
import google_auth_oauthlib.flow
import googleapiclient.discovery
from youtube_api import YoutubeDataApi
from youtube_api import youtube_api_utils
import youtube_api.parsers as parser
from sklearn.metrics import ndcg_score
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np

#Pull Excel Sheet
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def pull_excelRanks():
    wb = load_workbook("NDCG_Rubric.xlsx")
    sheet = wb.active
    true_scores = []
    true_vids = []
    you_scores = []
    you_vids = []
    rec_scores = []
    rec_vids = []
    for row in range(2, 12):
        if row <= 6:
            true_scores.append(int(sheet.cell(row=row, column=2).value))
            true_vids.append(str(sheet.cell(row=row, column=1).value))
            you_scores.append(int(sheet.cell(row=row, column=2).value))
            you_vids.append(str(sheet.cell(row=row, column=1).value))
        else:
            true_scores.append(int(sheet.cell(row=row, column=2).value))
            true_vids.append(str(sheet.cell(row=row, column=1).value))
            rec_scores.append(sheet.cell(row=row, column=2).value)
            rec_vids.append(str(sheet.cell(row=row, column=1).value))

    you_total = zip(you_scores, you_vids)
    rec_total = zip(rec_scores, rec_vids)
    true_total = zip(true_scores, true_vids)
    return true_total, rec_total, you_total
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~


#Eval NDCG
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def calc_NDCG(true_vids, you_vids, rec_vids, true_scores, rec_scores, you_scores):

    true_relevance = np.asarray([true_scores])
    youtube_NDCG = ndcg_score(true_relevance, np.asarray([you_scores]), k=5)
    recommenter_NDCG = ndcg_score(true_relevance, np.asarray([rec_scores]), k=5)
    with open("NDCG.txt", "w") as result:
        print(f"Ground Truth Videos  " + str(true_vids), file=result)
        print(f"Youtube Videos  " + str(you_vids), file=result)
        print(f"Recommenter Videos  " + str(rec_vids), file=result)
        print(f"Ground Truth  "+ str(true_scores), file=result)
        print(f"Youtube Ranked  "+ str(you_scores), file=result)
        print(f"Recommenter Ranked  " + str(rec_scores), file=result)

        print(f"YouTube NDCG " + str(youtube_NDCG), file=result)
        print(f"Recommenter NDCG " + str(recommenter_NDCG), file=result)
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

if __name__ == "__main__":
    true_total, rec_total, you_total = pull_excelRanks()
    sorted_true = sorted(true_total, key=lambda x: x[0], reverse= True)
    true_scores = [list(s) for s in zip(*sorted_true)][0][:5]
    true_Vids = [list(s) for s in zip(*sorted_true)][1][:5]

    you_scores = []
    you_vids = []
    for score,vid in you_total:
        you_vids.append(vid)
        if vid not in true_Vids:
            score = 0
        you_scores.append(score)

    rec_scores = []
    rec_vids = []
    for score, vid in rec_total:
        rec_vids.append(vid)
        if vid not in true_Vids:
            score = 0
        rec_scores.append(score)

    calc_NDCG(true_Vids, you_vids, rec_vids, true_scores, rec_scores, you_scores)


