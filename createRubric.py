import os
import google_auth_oauthlib.flow
import googleapiclient.discovery
from youtube_api import YoutubeDataApi
import youtube_api.parsers as parser
from openpyxl import Workbook
import json

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

#FILL IN API KEY BELOW
youtube = YoutubeDataApi('')
youtubeRecommended = []
recommenterRecommended = []
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

#Pull top K videos from youtube API
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def youtubeRecommender(ID, listy):
    videoInfo = youtube.get_recommended_videos( video_id= ID, max_results=5, parser= parser.raw_json )
    for dict in videoInfo:
        for id, value in dict["id"].items():
           if id == 'videoId':
               listy.append(value)
    return listy
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

#Pull Top K videos from Recommenter
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def recommenterRecommender(ID, listr):
    vids = []
    scores = []
    comment_weight = 0
    trans_weight = 0
    #Insert JSON FILE NAME HERE
    with open("recommenter_related_0.3.json") as json_file:
        rec_data= json.load(json_file)
        results = rec_data[ID]
        for key, value in results.items():
            vids.append(key)
            for feature, weight in value.items():
                if feature == "comment":
                    comment_weight = weight
                elif feature == "transcript":
                    trans_weight = weight

            totalWeight = trans_weight + comment_weight
            scores.append(totalWeight)

        rec_total = zip(scores, vids)
        sorted_rec = sorted(rec_total, key=lambda x: x[0], reverse=True)
        true_Vids = [list(s) for s in zip(*sorted_rec)][1][:5]



    return true_Vids

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

#Fill Excel Sheet
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
def fill_Excell(inputURL, youRanks, recRanks):
    #UPDATE FILE name if making several copies
    filename = "NDCG_Rubric.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.column_dimensions['A'].width = 60
    sheet.cell(row=1, column=1).value = "Comparison Video =  " + inputURL
    sheet.cell(row=1, column=2).value = "Rank 1 - 5"
    for row in range(2, 12):
        if row <= 6:
            sheet.cell(row=row, column=1).hyperlink = "https://www.youtube.com/watch?v=" + youRanks[row - 2]
            sheet.cell(row=row, column=1).style = "Hyperlink"
        else:
            ind = row - 6
            sheet.cell(row=row, column=1).hyperlink = "https://www.youtube.com/watch?v=" + recRanks[ind-1]
            sheet.cell(row=row, column=1).style = "Hyperlink"
    workbook.save(filename=filename)
#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

def mainInput(inputURL):
#have to run first once to populate excel, then comment out until the input line to pull and uncomment below input line
    Video_ID = inputURL.partition('v=')[2]
    youtubeRecommended = youtubeRecommender(Video_ID, [])
    recommenterRecommended = recommenterRecommender(Video_ID, [])
    fill_Excell(inputURL, youtubeRecommended, recommenterRecommended)

if __name__ == "__main__":
    #Update seed video for each run
    mainInput('https://www.youtube.com/watch?v=No8-mBek3rs')